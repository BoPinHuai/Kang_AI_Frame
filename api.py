import os
# 离线友好：必须在 import chromadb / sentence-transformers 之前设置
os.environ.setdefault("ANONYMIZED_TELEMETRY", "False")
os.environ.setdefault("CHROMA_TELEMETRY_DISABLED", "1")
os.environ.setdefault("HF_HUB_DISABLE_TELEMETRY", "1")

import json
import uuid
import shutil
from pathlib import Path
from datetime import datetime, timedelta

import threading
from typing import Optional
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from loader import SUPPORTED
from indexer import build_index
from chat import ask_stream
from retriever import reset_collection
from settings import load_settings, save_settings
from config import (
    load_config, save_config, public_view, merge_update,
    get_provider_by_name,
)
from providers import ProviderError

# ── 路径配置 ────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
DB_PATH     = str(BASE_DIR / "db")
META_PATH   = str(BASE_DIR / "db_meta.json")
KB_DESC_PATH = BASE_DIR / "kb_desc.json"   # 知识库简述（folder → 描述）
LIBRARY_DIR = BASE_DIR / "library"
HISTORY_DIR = BASE_DIR / "history"
STATIC_DIR  = BASE_DIR / "static"

LIBRARY_DIR.mkdir(exist_ok=True)
HISTORY_DIR.mkdir(exist_ok=True)
STATIC_DIR.mkdir(exist_ok=True)

# SUPPORTED 从 loader 统一导入（.pdf .docx .md .xlsx .csv）


def _warmup():
    """后台预热：embedding 模型 + ChromaDB；仅 Ollama provider 时才预热 LLM。"""
    from embedder import get_model
    from retriever import _get_collection
    from config import get_active_provider, load_config
    try:
        get_model()
        print("[预热] embedding 模型就绪")
    except Exception as e:
        print(f"[预热] embedding 加载失败：{e}")
    try:
        _get_collection(DB_PATH)
        print("[预热] ChromaDB 就绪")
    except Exception as e:
        print(f"[预热] ChromaDB 加载失败（可能索引尚未建立）：{e}")
    # 只对本地 Ollama 做 LLM 预热；外部 API 不需要也不应该消耗 token
    try:
        cfg = load_config()
        active_name = cfg.get("active", "")
        active_type = cfg.get("providers", {}).get(active_name, {}).get("type", "ollama")
        if active_type == "ollama":
            provider = get_active_provider()
            for _ in provider.stream_chat([{"role": "user", "content": "hi"}]):
                break
            print("[预热] Ollama LLM 就绪")
        else:
            print(f"[预热] 跳过 LLM 预热（外部 API: {active_type}/{active_name}）")
    except Exception as e:
        print(f"[预热] LLM 预热失败（首次查询时再加载）：{e}")


app = FastAPI()

# 模块加载时立即预热，不等 startup event
threading.Thread(target=_warmup, daemon=True).start()


# ── 启动：后台增量更新索引 ────────────────────────────────
@app.on_event("startup")
async def startup_event():
    files = [f for f in LIBRARY_DIR.rglob("*")
             if f.is_file() and f.suffix.lower() in SUPPORTED and not f.name.startswith("~$")]
    if files:
        force = not Path(META_PATH).exists()
        action = "强制全量重建（首次迁移）" if force else f"增量检查（共 {len(files)} 个文件）"
        print(f"[启动] {action}（后台进行）...")
        threading.Thread(
            target=build_index,
            args=(str(LIBRARY_DIR), DB_PATH, META_PATH),
            kwargs={"force": force},
            daemon=True,
        ).start()
    else:
        print("[启动] library/ 为空，跳过索引")


# ── 数据模型 ─────────────────────────────────────────
class ChatRequest(BaseModel):
    question: str
    conversation_id: Optional[str] = None
    scope: Optional[str] = None   # 检索范围，如 "产品手册" 或 "产品手册/说明书.pdf"


class MkdirRequest(BaseModel):
    path: str      # 相对 library/ 的路径，如 "产品手册/接口文档"


class SaveMessageRequest(BaseModel):
    conv_id: Optional[str] = None
    question: str
    answer: str
    sources: list = []


class ConfigUpdateRequest(BaseModel):
    active: Optional[str] = None
    providers: Optional[dict] = None


class TestConnectionRequest(BaseModel):
    name: Optional[str] = None     # 按名字测试已存在的 provider
    profile: Optional[dict] = None  # 或临时传一个配置过来测试（不入库）


# ── 工具函数 ─────────────────────────────────────────
def save_history(conv_id: Optional[str], question: str, answer: str, sources: list) -> str:
    conv_id = conv_id or str(uuid.uuid4())
    history_file = HISTORY_DIR / f"{conv_id}.json"

    if history_file.exists():
        with open(history_file, encoding="utf-8") as f:
            history = json.load(f)
    else:
        title = question.replace("\n", " ")[:28]
        history = {"id": conv_id, "title": title, "created": datetime.now().isoformat(), "messages": []}

    now = datetime.now().isoformat()
    history["messages"].append({"role": "user", "content": question, "timestamp": now})
    history["messages"].append({"role": "assistant", "content": answer, "sources": sources, "timestamp": now})

    with open(history_file, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)
    return conv_id


def build_library_tree(folder: Path, root: Path) -> dict:
    """递归构建文件树 JSON"""
    children = []
    try:
        items = sorted(folder.iterdir(),
                       key=lambda x: (0 if x.is_dir() else 1, x.name.lower()))
    except PermissionError:
        items = []
    for item in items:
        if item.name.startswith("~$") or item.name.startswith("."):
            continue
        if item.is_dir():
            children.append(build_library_tree(item, root))
        elif item.is_file() and item.suffix.lower() in SUPPORTED:
            rel = str(item.relative_to(root)).replace("\\", "/")
            children.append({
                "name": item.name,
                "type": "file",
                "path": rel,
                "size": f"{round(item.stat().st_size / 1024, 1)} KB",
                "modified": datetime.fromtimestamp(item.stat().st_mtime).strftime("%Y-%m-%d"),
            })
    rel_path = str(folder.relative_to(root)).replace("\\", "/") if folder != root else ""
    return {"name": folder.name, "type": "folder", "path": rel_path, "children": children}


# ── 后台任务（索引进度） ───────────────────────────
_jobs: dict[str, dict] = {}
_jobs_lock = threading.Lock()


def _new_job() -> str:
    job_id = uuid.uuid4().hex[:12]
    with _jobs_lock:
        _jobs[job_id] = {
            "status": "running",
            "current": 0,
            "total": 0,
            "message": "准备中…",
            "result": None,
            "error": None,
        }
    return job_id


def _update_job(job_id: str, **fields):
    with _jobs_lock:
        if job_id in _jobs:
            _jobs[job_id].update(fields)


def _run_index_job(job_id: str, force: bool = False, post_message: str = None):
    """后台任务：跑 build_index 并更新 _jobs[job_id]。"""
    def cb(cur, tot, msg):
        _update_job(job_id, current=cur, total=tot, message=msg)
    try:
        indexed, skipped = build_index(str(LIBRARY_DIR), DB_PATH, META_PATH,
                                        force=force, progress_cb=cb)
        reset_collection()
        msg = post_message or f"索引完成：新增/更新 {indexed}，跳过 {skipped}"
        _update_job(job_id, status="done", message=msg,
                    result={"indexed": indexed, "skipped": skipped})
    except Exception as e:
        _update_job(job_id, status="error", error=str(e), message=f"失败：{e}")


# ── 流式对话 ─────────────────────────────────────────
@app.post("/api/chat/stream")
async def chat_stream(request: ChatRequest):
    history = []
    if request.conversation_id:
        hf = HISTORY_DIR / f"{request.conversation_id}.json"
        if hf.exists():
            with open(hf, encoding="utf-8") as f:
                conv = json.load(f)
            past = conv.get("messages", [])[-6:]
            history = [{"role": m["role"], "content": m["content"]} for m in past]

    def generate():
        full_answer, sources = [], []
        try:
            for event_type, data in ask_stream(request.question, DB_PATH,
                                                history=history, scope=request.scope):
                if event_type == "sources":
                    sources = data
                    yield f"data: {json.dumps({'type': 'sources', 'sources': data}, ensure_ascii=False)}\n\n"
                elif event_type == "chunk":
                    full_answer.append(data)
                    yield f"data: {json.dumps({'type': 'chunk', 'content': data}, ensure_ascii=False)}\n\n"
                elif event_type == "loading":
                    yield f"data: {json.dumps({'type': 'loading', 'message': data}, ensure_ascii=False)}\n\n"
                elif event_type == "error":
                    yield f"data: {json.dumps({'type': 'error', 'message': data}, ensure_ascii=False)}\n\n"
                    return
                elif event_type == "done":
                    answer = "".join(full_answer)
                    conv_id = save_history(request.conversation_id, request.question, answer, sources)
                    yield f"data: {json.dumps({'type': 'done', 'conversation_id': conv_id}, ensure_ascii=False)}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'type': 'error', 'message': f'服务器错误：{str(e)}'}, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── 仪表盘统计 ────────────────────────────────────────
def _human_size(nbytes: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if nbytes < 1024 or unit == "GB":
            return f"{nbytes:.1f} {unit}" if unit != "B" else f"{nbytes} B"
        nbytes /= 1024
    return f"{nbytes:.1f} GB"


def compute_stats() -> dict:
    """实时从磁盘计算统计：不依赖任何额外存储。"""
    # 1) 扫 library：文档数 / 大小 / 类型占比 / 各知识库（顶层文件夹）
    files = [f for f in LIBRARY_DIR.rglob("*")
             if f.is_file() and f.suffix.lower() in SUPPORTED
             and not f.name.startswith("~$") and not f.name.startswith(".")]
    total_bytes = sum(f.stat().st_size for f in files)
    by_type: dict[str, int] = {}
    kb: dict[str, dict] = {}
    for f in files:
        ext = f.suffix.lower().lstrip(".").upper()
        by_type[ext] = by_type.get(ext, 0) + 1
        rel = f.relative_to(LIBRARY_DIR)
        top = rel.parts[0] if len(rel.parts) > 1 else "（根目录）"
        k = kb.setdefault(top, {"name": top, "docs": 0, "bytes": 0})
        k["docs"] += 1
        k["bytes"] += f.stat().st_size

    # 2) 已索引数（解析成功）来自 db_meta.json
    indexed = 0
    try:
        with open(META_PATH, encoding="utf-8") as fp:
            meta = json.load(fp)
        indexed = len([k for k in meta if not k.startswith("__")])
    except Exception:
        pass

    # 3) 问答统计 + 近 14 天趋势，遍历 history/*.json
    from collections import defaultdict
    conv_count = 0
    msg_count = 0
    day_conv: dict[str, set] = defaultdict(set)
    day_msg: dict[str, int] = defaultdict(int)
    for hf in HISTORY_DIR.glob("*.json"):
        try:
            with open(hf, encoding="utf-8") as fp:
                conv = json.load(fp)
        except Exception:
            continue
        conv_count += 1
        for m in conv.get("messages", []):
            if m.get("role") != "user":
                continue
            msg_count += 1
            ts = (m.get("timestamp") or "")[:10]
            if ts:
                day_conv[ts].add(conv.get("id", hf.stem))
                day_msg[ts] += 1

    # 近 14 天序列（无数据的日期补 0）
    today = datetime.now().date()
    trend = []
    for i in range(13, -1, -1):
        d = today - timedelta(days=i)
        key = d.isoformat()
        trend.append({"date": d.strftime("%m-%d"),
                      "convs": len(day_conv.get(key, set())),
                      "msgs": day_msg.get(key, 0)})

    kb_list = sorted(
        [{"name": v["name"], "docs": v["docs"], "size": _human_size(v["bytes"])}
         for v in kb.values()],
        key=lambda x: -x["docs"],
    )

    return {
        "doc_count": len(files),
        "total_size": _human_size(total_bytes),
        "indexed_count": indexed,
        "pending_count": max(0, len(files) - indexed),
        "kb_count": len(kb),
        "by_type": sorted(
            [{"ext": k, "count": v} for k, v in by_type.items()],
            key=lambda x: -x["count"]),
        "kb_list": kb_list,
        "conv_count": conv_count,
        "msg_count": msg_count,
        "trend": trend,
    }


@app.get("/api/stats")
async def get_stats():
    return compute_stats()


@app.get("/api/activity")
async def get_activity():
    """侧栏知识库 tab：最近索引动态 + 最近一次问答。"""
    # 最近索引的文件（按 mtime 取前 5）
    recent_files = []
    try:
        with open(META_PATH, encoding="utf-8") as f:
            meta = json.load(f)
        pairs = [(k, v.get("mtime", 0)) for k, v in meta.items() if not k.startswith("__")]
        pairs.sort(key=lambda x: -x[1])
        for rel, mtime in pairs[:5]:
            recent_files.append({
                "path": rel,
                "name": Path(rel).name,
                "modified": datetime.fromtimestamp(mtime).strftime("%m-%d %H:%M"),
            })
    except Exception:
        pass
    # 最近一次问答（取最新的 history 文件的最后一问一答）
    latest_qa = None
    try:
        files = sorted(HISTORY_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True)
        if files:
            with open(files[0], encoding="utf-8") as f:
                conv = json.load(f)
            msgs = conv.get("messages", [])
            q = next((m["content"] for m in reversed(msgs) if m["role"] == "user"), None)
            a = next((m["content"] for m in reversed(msgs) if m["role"] == "assistant"), None)
            srcs = next((m.get("sources", []) for m in reversed(msgs) if m["role"] == "assistant"), [])
            if q and a:
                latest_qa = {
                    "question": q[:80] + ("…" if len(q) > 80 else ""),
                    "answer": a[:120] + ("…" if len(a) > 120 else ""),
                    "sources": list({s["source"] for s in srcs})[:3],
                }
    except Exception:
        pass
    # 索引概况
    indexed = len([k for k in (meta if "meta" in dir() else {}) if not k.startswith("__")])
    return {"indexed": indexed, "recent_files": recent_files, "latest_qa": latest_qa}


# ── 知识库（顶层文件夹的卡片视图）────────────────────────
def _load_kb_desc() -> dict:
    try:
        with open(KB_DESC_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_kb_desc(d: dict):
    with open(KB_DESC_PATH, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=2)


def _kb_meta(d: dict, path: str):
    """读取某知识库的 (desc, icon)，兼容旧版纯字符串存储。"""
    e = d.get(path)
    if isinstance(e, dict):
        return e.get("desc", ""), e.get("icon", "")
    if isinstance(e, str):
        return e, ""
    return "", ""


@app.get("/api/kb")
async def list_kb():
    """列出每个知识库（library 顶层文件夹）+ 根目录散落文件，含文档数/大小/简述/图标。"""
    meta = _load_kb_desc()
    kb: dict[str, dict] = {}
    root_docs, root_bytes = 0, 0
    for f in LIBRARY_DIR.rglob("*"):
        if not (f.is_file() and f.suffix.lower() in SUPPORTED):
            continue
        if f.name.startswith("~$") or f.name.startswith("."):
            continue
        rel = f.relative_to(LIBRARY_DIR)
        size = f.stat().st_size
        if len(rel.parts) > 1:
            top = rel.parts[0]
            k = kb.setdefault(top, {"name": top, "path": top, "docs": 0, "bytes": 0})
            k["docs"] += 1
            k["bytes"] += size
        else:
            root_docs += 1
            root_bytes += size
    items = []
    for v in sorted(kb.values(), key=lambda x: -x["docs"]):
        desc, icon = _kb_meta(meta, v["path"])
        items.append({"name": v["name"], "path": v["path"], "docs": v["docs"],
                      "size": _human_size(v["bytes"]), "desc": desc, "icon": icon})
    return {"kbs": items,
            "root": {"docs": root_docs, "size": _human_size(root_bytes)}}


class KbMetaRequest(BaseModel):
    path: str
    desc: str = ""
    icon: str = ""


@app.put("/api/kb/meta")
async def set_kb_meta(req: KbMetaRequest):
    """保存知识库的简述 + 自定义图标（base64 data URL）。"""
    d = _load_kb_desc()
    path = req.path.strip("/")
    desc = req.desc.strip()[:120]
    icon = req.icon or ""
    if desc or icon:
        d[path] = {"desc": desc, "icon": icon}
    else:
        d.pop(path, None)
    _save_kb_desc(d)
    return {"ok": True, "path": path}


class KbRenameRequest(BaseModel):
    path: str
    new_name: str


@app.post("/api/kb/rename")
async def rename_kb(req: KbRenameRequest, bg: BackgroundTasks):
    """重命名知识库（顶层文件夹），迁移其简述/图标，并后台重建索引以更新来源路径。"""
    old = req.path.strip("/")
    new = req.new_name.strip().replace("/", "").replace("\\", "").replace("..", "").strip()
    if not new:
        raise HTTPException(status_code=400, detail="名称不能为空")
    src = LIBRARY_DIR / old
    dst = LIBRARY_DIR / new
    if not src.is_dir():
        raise HTTPException(status_code=404, detail="知识库不存在")
    if dst.exists():
        raise HTTPException(status_code=400, detail=f"「{new}」已存在")
    src.rename(dst)
    # 迁移简述/图标
    d = _load_kb_desc()
    if old in d:
        d[new] = d.pop(old)
        _save_kb_desc(d)
    # 路径变了 → 后台增量重建（旧路径块移除、新路径块加入）
    job_id = _new_job()
    bg.add_task(_run_index_job, job_id, False, f"已重命名为「{new}」，索引已更新")
    return {"ok": True, "new_name": new, "job_id": job_id}


# ── 文件预览 ─────────────────────────────────────────
@app.get("/api/preview/{path:path}")
async def preview_file(path: str):
    """返回文件预览数据：MD/TXT→文本, XLSX/CSV→表格, DOCX→文本, PDF→重定向到原始文件。"""
    from fastapi.responses import FileResponse
    clean = path.lstrip("/").replace("..", "").strip()
    abs_path = LIBRARY_DIR / clean
    if not abs_path.exists() or not abs_path.is_file():
        raise HTTPException(status_code=404, detail="文件不存在")
    # 安全检查：必须在 library 内
    try:
        abs_path.relative_to(LIBRARY_DIR)
    except ValueError:
        raise HTTPException(status_code=403, detail="访问被拒绝")

    suffix = abs_path.suffix.lower()

    if suffix == ".pdf":
        return FileResponse(str(abs_path), media_type="application/pdf",
                            headers={"Content-Disposition": f"inline; filename=\"{abs_path.name}\""})

    if suffix in (".md", ".txt"):
        text = abs_path.read_text(encoding="utf-8", errors="ignore")
        return {"type": "markdown", "content": text, "name": abs_path.name}

    if suffix == ".docx":
        from loader import _seg_docx
        segs = _seg_docx(str(abs_path))
        text = "\n\n".join(s["text"] for s in segs if s["text"].strip())
        return {"type": "text", "content": text, "name": abs_path.name}

    if suffix == ".csv":
        import csv, io
        rows = []
        with open(abs_path, newline="", encoding="utf-8-sig", errors="ignore") as f:
            for r in csv.reader(f):
                rows.append(r)
                if len(rows) > 200:
                    break
        return {"type": "table", "rows": rows, "name": abs_path.name}

    if suffix == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(str(abs_path), read_only=True, data_only=True)
        sheets = {}
        for ws in wb.worksheets:
            rows = []
            for r in ws.iter_rows(values_only=True):
                rows.append(["" if c is None else str(c) for c in r])
                if len(rows) > 200:
                    break
            if rows:
                sheets[ws.title] = rows
        wb.close()
        return {"type": "xlsx", "sheets": sheets, "name": abs_path.name}

    raise HTTPException(status_code=415, detail=f"不支持预览该格式：{suffix}")


# ── 文件库 ───────────────────────────────────────────
@app.get("/api/library")
async def get_library():
    """返回 library/ 完整文件树"""
    return build_library_tree(LIBRARY_DIR, LIBRARY_DIR)


@app.post("/api/library/mkdir")
async def make_directory(request: MkdirRequest):
    """在 library/ 下新建文件夹（支持多级）"""
    # 防止路径穿越
    clean = request.path.strip("/").replace("..", "").strip()
    if not clean:
        raise HTTPException(status_code=400, detail="路径不能为空")
    target = LIBRARY_DIR / clean
    if target.exists():
        raise HTTPException(status_code=400, detail="文件夹已存在")
    target.mkdir(parents=True, exist_ok=True)
    return {"message": f"文件夹「{clean}」创建成功", "path": clean}


@app.post("/api/upload")
async def upload_document(
    bg: BackgroundTasks,
    file: UploadFile = File(...),
    folder: str = Form(default=""),
):
    """上传文件 → 立即返回 job_id，索引在后台跑。前端轮询 /api/jobs/{id} 看进度。"""
    suffix = Path(file.filename).suffix.lower()
    if suffix not in SUPPORTED:
        raise HTTPException(status_code=400, detail="仅支持 PDF、Word(.docx)、Markdown(.md)、Excel(.xlsx)、CSV(.csv)")

    if folder:
        clean_folder = folder.strip("/").replace("..", "").strip()
        dest_dir = LIBRARY_DIR / clean_folder
    else:
        dest_dir = LIBRARY_DIR
    dest_dir.mkdir(parents=True, exist_ok=True)

    dest = dest_dir / file.filename
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)

    rel = str(dest.relative_to(LIBRARY_DIR)).replace("\\", "/")
    location = folder if folder else "根目录"

    job_id = _new_job()
    post_msg = f"「{file.filename}」已上传到「{location}」，索引已更新"
    bg.add_task(_run_index_job, job_id, False, post_msg)

    return {"job_id": job_id, "filename": file.filename, "location": location, "path": rel}


@app.post("/api/rebuild")
async def rebuild_index(bg: BackgroundTasks, force: bool = False):
    """重建索引 → 立即返回 job_id，索引在后台跑。"""
    files = [f for f in LIBRARY_DIR.rglob("*")
             if f.is_file() and f.suffix.lower() in SUPPORTED and not f.name.startswith("~$")]
    if not files:
        raise HTTPException(status_code=400, detail="library/ 文件夹为空，请先上传文档")
    job_id = _new_job()
    bg.add_task(_run_index_job, job_id, force, None)
    return {"job_id": job_id, "total_files": len(files)}


@app.get("/api/jobs/{job_id}")
async def get_job(job_id: str):
    with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="job 不存在或已过期")
        return dict(job)


# ── 模型状态 ─────────────────────────────────────────
@app.get("/api/status")
async def get_status():
    cfg = load_config()
    active_name = cfg.get("active", "")
    active_type = cfg.get("providers", {}).get(active_name, {}).get("type", "ollama")

    if active_type != "ollama":
        # 外部 API（DeepSeek / OpenAI / Kimi …）：配置存在即就绪，不调 Ollama
        return {"status": "ready", "models": [], "type": active_type, "active": active_name}

    # Ollama：尝试列出本地模型
    try:
        import ollama
        result = ollama.list()
        names = [m["model"] for m in result.get("models", [])]
        return {"status": "ready", "models": names, "type": "ollama", "active": active_name}
    except Exception:
        return {"status": "unavailable", "models": [], "type": "ollama", "active": active_name}


# ── 历史对话 ─────────────────────────────────────────
@app.get("/api/history")
async def list_history():
    items = []
    for f in sorted(HISTORY_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
        with open(f, encoding="utf-8") as fp:
            d = json.load(fp)
            items.append({"id": d["id"], "title": d["title"], "created": d["created"]})
    return items


@app.get("/api/history/{conv_id}")
async def get_history(conv_id: str):
    f = HISTORY_DIR / f"{conv_id}.json"
    if not f.exists():
        raise HTTPException(status_code=404, detail="对话不存在")
    with open(f, encoding="utf-8") as fp:
        return json.load(fp)


@app.delete("/api/history/{conv_id}")
async def delete_history(conv_id: str):
    f = HISTORY_DIR / f"{conv_id}.json"
    if f.exists():
        f.unlink()
    return {"message": "对话已删除"}


@app.delete("/api/history/{conv_id}/last")
async def delete_last_exchange(conv_id: str):
    f = HISTORY_DIR / f"{conv_id}.json"
    if not f.exists():
        raise HTTPException(status_code=404, detail="对话不存在")
    with open(f, encoding="utf-8") as fp:
        conv = json.load(fp)
    msgs = conv.get("messages", [])
    while msgs and msgs[-1]["role"] == "assistant":
        msgs.pop()
    while msgs and msgs[-1]["role"] == "user":
        msgs.pop()
    conv["messages"] = msgs
    with open(f, "w", encoding="utf-8") as fp:
        json.dump(conv, fp, ensure_ascii=False, indent=2)
    return {"message": "已删除最后一轮"}


@app.post("/api/save-message")
async def save_message(request: SaveMessageRequest):
    if not request.answer.strip():
        return {"conversation_id": request.conv_id}
    conv_id = save_history(request.conv_id, request.question, request.answer, request.sources)
    return {"conversation_id": conv_id}


# ── 模型配置 ─────────────────────────────────────────
@app.get("/api/config")
async def get_config():
    """返回当前 config（api_key 已遮蔽）。"""
    return public_view(load_config())


@app.put("/api/config")
async def update_config(request: ConfigUpdateRequest):
    """更新 config。空 api_key 字段会保留原值，避免前端遮蔽后误清空。"""
    cfg = load_config()
    new_cfg = merge_update(cfg, request.dict(exclude_unset=True))
    save_config(new_cfg)
    return public_view(new_cfg)


@app.post("/api/config/test")
async def test_connection(request: TestConnectionRequest):
    """测试 provider 连接。可按 name 测试已保存的，或传 profile 临时测试。"""
    try:
        if request.profile:
            from providers import build_provider
            # 若 profile 没带 api_key，但同名 provider 已有 key，则借用
            profile = dict(request.profile)
            if not profile.get("api_key"):
                cfg = load_config()
                name = request.name or profile.get("name")
                if name and name in cfg["providers"]:
                    profile["api_key"] = cfg["providers"][name].get("api_key", "")
            provider = build_provider(profile)
        elif request.name:
            provider = get_provider_by_name(request.name)
        else:
            raise HTTPException(status_code=400, detail="需提供 name 或 profile")
        ok, message = provider.test()
        return {"ok": ok, "message": message}
    except ProviderError as e:
        return {"ok": False, "message": str(e)}
    except HTTPException:
        raise
    except Exception as e:
        return {"ok": False, "message": f"测试失败：{e}"}


# ── 参数设置 ─────────────────────────────────────────
@app.get("/api/settings")
async def get_settings():
    s = load_settings()
    # 密码不暴露给前端（只告知是否启用）
    s.pop("lock_password", None)
    return s


@app.get("/api/lock-status")
async def lock_status():
    """告诉前端当前是否需要密码。不暴露密码本身。"""
    s = load_settings()
    return {"locked": s.get("lock_enabled", False) and bool(s.get("lock_password", ""))}


class UnlockRequest(BaseModel):
    password: str


@app.post("/api/unlock")
async def unlock(req: UnlockRequest):
    """验证密码，服务端对比，前端不会拿到密码明文。"""
    s = load_settings()
    if not s.get("lock_enabled") or not s.get("lock_password"):
        return {"ok": True}
    return {"ok": req.password == s["lock_password"]}


@app.put("/api/settings")
async def update_settings(request: Request):
    data = await request.json()
    # 密码字段允许通过 PUT 设置（来自个人资料页）
    result = save_settings(data)
    result.pop("lock_password", None)   # 响应不暴露密码
    return result


# ── 静态文件 ─────────────────────────────────────────
app.mount("/", StaticFiles(directory=str(STATIC_DIR), html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("api:app", host="127.0.0.1", port=8000, reload=False)
